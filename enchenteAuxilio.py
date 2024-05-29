#===========================================================================================================================================
# -*- coding: utf-8 -*-  # Arquivo: enchentesAuxilio.py
# Versão: 1.0
# Última alteração: 24/05/2024
# Propósito: Gerar planilha consolidada para auxilio enchente
# Autor: Roberto Edgar Geiss 
# Copyright: PMNH
# Produto: 
# Observacoes:  
# Parametros: 
# Detalhes especificos: 
#===========================================================================================================================================
from openpyxl import Workbook
from openpyxl  import load_workbook
import os, sys, time, re
import pandas as pd

# inicializa variaveis
arq_exec = ""
arq_linhas = ""
int_Linhas = 0
int_Lidas = 0
cel_cepc = ""
cel_namef = ""
cel_cepf = ""
cel_cepc = ""
nro_end = 0
cod_familia = 0
nro_familia = 0

linha = ""
apenda = False
lista = ()

colunaA = ""
colunaB = ""
colunaC = ""
colunaD = ""
colunaE = ""
colunaF = ""
colunaG = ""
colunaH = ""
colunaI = ""
colunaL = ""
colunaM = ""
colunaO = ""
# dados dos familiares
colunaFC1 = ""
colunaFN1 = ""
colunaFC2 = ""
colunaFN2 = ""
colunaFC3 = ""
colunaFN3 = ""
colunaFC4 = ""
colunaFN4 = ""
colunaFC5 = ""
colunaFN5 = ""
colunaFC6 = ""
colunaFN6 = ""
colunaFC7 = ""
colunaFN7 = ""
colunaFC8 = ""
colunaFN8 = ""
colunaFC9 = ""
colunaFN9 = ""
colunaFC10 = ""
colunaFN10 = ""

strLinha = ""
strLinhaCabec = "CPF_RESPONSAVEL" + ';' + "NOME_RESPONSAVEL" + ';' + "LOGRADOURO" + ';' + "NUMERO" + ';' + "COMPLEMENTO" + ';' + "BAIRRO" + ';' + "CEP" + ';' + "TELEFONE_RESPONSAVEL" + ';' + \
                "CPF_MEMBRO1" + ';' + "NOME_MEMBRO1" + ';' + "CPF_MEMBRO2" + ';' + "NOME_MEMBRO2" + ';' + "CPF_MEMBRO3" + ';' + "NOME_MEMBRO3" + ';' + "CPF_MEMBRO4" + ';' + "NOME_MEMBRO4" + \
                 ';' + "CPF_MEMBRO5" + ';' + "NOME_MEMBRO5" + ';' + "CPF_MEMBRO6" + ';' + "NOME_MEMBRO6" + ';' + "CPF_MEMBRO7" + ';' + "NOME_MEMBRO7" + ';' + "CPF_MEMBRO8" + ';' + "NOME_MEMBRO8" + \
                 ';' + "CPF_MEMBRO9" + ';' + "NOME_MEMBRO9" + ';' + "CPF_MEMBRO10" + ';' + "NOME_MEMBRO10"
strQuote = "'"
strDblQuote = '"'

# Funções 
#--------
def validaCEP(cep, nro):
    
    if nro:
        nro =  int(nro) 
    else:
        nro = 0
        
    nro_ini = 0
    nro_fim = 0

    apenda = False                
    
    for row in range(2, sheetc.max_row + 1):
        
        for column in "ABC":
            
            if column == "A":
                cell_name = "{}{}".format(column, row)   
                cel_cep = sheetc[cell_name].value     
                
            if column == "B":
                cell_name = "{}{}".format(column, row)   
                nro_ini = sheetc[cell_name].value
                nro_ini = int(nro_ini)

            if column == "C":
                cell_name = "{}{}".format(column, row)   
                nro_fim = sheetc[cell_name].value
                nro_fim = int(nro_fim) 

                if cel_cep == cep:
                    # Toda a extensão da rua    
                    if nro_ini == 0 and nro_fim == 0:
                        print(cel_cep, cel_cepf, sep=" - ")
                        apenda = True   
                    
                    # Do inicio ate nro da rua
                    if nro_ini == 0 and nro_fim > 0:
                        if nro <= nro_fim:
                            print(cel_cep, cel_cepf, sep=" - ")
                            apenda = True
                        
                    # De nro ate o fim da rua
                    if nro_ini > 0 and nro_fim == 0:
                        if nro >= nro_ini:
                            print(cel_cep, cel_cepf, sep=" - ")
                            apenda = True
                        
                    # Do nro exato
                    if nro_ini == nro_fim:
                        if nro_ini == nro:                            
                            print(cel_cep, cel_cepf, sep=" - ")
                            apenda = True

    return apenda 

# Abre arquivo de log 
#--------------------
try: 
    arq_exec = open("logExecucao.txt", "w")
    arq_linhas = open("linhasCSV.csv", "w")

except IOError: 
    print ("Erro abrindo arquivo log: "), arq_exec 
    raise
    sys.exit()

localTime = time.asctime(time.localtime(time.time()))  
strHora = "-- Gerado em : " +  localTime  + "\n" 
arq_exec.write(strHora)

# Classifica as planilhas com pandas
#-----------------------------------
try: 

    print ("Abrindo as planilhas")
    dff = pd.read_excel('familiasGov.xlsx')
    dfc = pd.read_excel('ceps1.xlsx')

    print ("Classificando as planilhas")
    dff.sort_values(by=['num_cep_logradouro_fam', 'cod_familiar_fam'], ascending=True, inplace=True, ignore_index=True)
    dfc.sort_values(by='CEPs', ascending=True, inplace=True, ignore_index=True)

except Exception:
    raise
    sys.exit()

# Abre planilha cadunico
wbf = load_workbook(filename="familiasGov.xlsx", data_only = True)
sheetf = wbf.active

# Abre planilha ceps
wbc = load_workbook(filename="ceps1.xlsx", data_only = True)
sheetc = wbc.active 

# Percorre as planilhas
#----------------------
print ("Filtrando as planilhas")
try:
    # grava cabecalho
    arq_linhas.write(strLinhaCabec  + '\n')  
    apenda = False 
    contador = 0
    cod_familia = 0
    
    for row in range(2, 100):              # sheetf.max_row + 1):
        int_Lidas = int_Lidas + 1
        
        print("Cod familia => linha")
        print(colunaA)  
        print(cod_familia)
        print(row)
        print(contador)
        # breakpoint()
        for column in "ABCDEFGHIJKLMNO":
            
            cell_namef = "{}{}".format(column, row)   
            celula = sheetf[cell_namef].value

            if celula == None: 
                celula = "" 
 
            if column == "A":     # codigo familia  
                colunaA = celula   
                cod_familia = celula
            if column == "B":     # bairro 
                colunaB = celula
            elif column == "C":   # tipo logradouro 
                colunaC = celula
            elif column == "D":   # denominacao logrdouro
                colunaD = celula
            elif column == "E":   # logradouro 
                colunaE = celula
            elif column == "F":   # nro logradouro
                colunaF = celula
                nro_end = celula 
            elif column == "G":   # complemento
                colunaG = celula
            elif column == "H":   # cep 
                colunaH = celula
                cel_cepf = celula    
                apenda = validaCEP(celula, nro_end)    
            elif column == "I":   # nro de pessoas na familia 
                colunaI = celula
                nro_familia = celula
            elif column == "J":   # ddd
                colunaJ = celula
            elif column == "K":   # telefone
                colunaK = celula
            elif column == "L":   # cod familia membro
                colunaL = celula
            elif column == "M":   # nome familiar
                colunaM = celula
            elif column == "N":   # parentesco familiar
                colunaN = celula
            elif column == "O":   # cpf familiar 
                colunaO = celula

 
        # breakpoint()
        if apenda == True:
            # breakpoint()
            # agrupa por familia 
            print("Cod familia => ")
            print(colunaA)  
            print(cod_familia)
            print(contador)
                  
            strLinha = str(colunaO) + ';' + strQuote + str(colunaM) + strQuote + ';' + strQuote + str(colunaC) + " " + str(colunaD) + " " + str(colunaE) + strQuote + ';' + \
                str(colunaF) + ';' + strQuote + str(colunaG) + strQuote + ';' + str(colunaB) + ';' + \
                str(colunaH) + ';' + str(colunaJ) + str(colunaK) 

            if cod_familia == colunaA:

                if contador > 0 and contador <= 10:
                    strLinha = strLinha + ';' + str(colunaO) + ';' + strQuote + str(colunaM)
                    contador = contador + 1

            else:            
                arq_linhas.write(strLinha  + '\n')  
                apenda = False 
                int_Linhas = int_Linhas + 1
                cod_familia = colunaA 
                contador = 0

except KeyError: 
    print ("Erro de atributo: "), arq_exec 
    print ("Linhas lidas => " + str(int_Linhas))  
    print ("Linhas lidas => " + str(int_Linhas))   
    arq_exec.close() 
    arq_linhas.close()
    raise
    sys.exit()

# Finalização
#------------
localTime = time.asctime( time.localtime(time.time()) )  
strHora = "-- Gerado em : " +  localTime  
print ("Linhas geradas => " + str(int_Linhas)) 
print ("Linhas lidas => " + str(int_Lidas))   

# -- fim do arquivo ----
